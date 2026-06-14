import base64
import io
import zipfile

import fitz
from docx import Document
from docx.oxml.ns import qn
from lxml import etree
from openpyxl import load_workbook


def extract_content(file_bytes: bytes, filename: str) -> list:
    ext = filename.rsplit(".", 1)[-1].lower()

    if ext == "pdf":
        return _extract_pdf(file_bytes)
    elif ext == "docx":
        return _extract_docx(file_bytes)
    elif ext == "xlsx":
        return _extract_xlsx(file_bytes)
    elif ext in ("png", "jpg", "jpeg", "gif", "webp"):
        return _extract_image(file_bytes, ext)
    elif ext == "txt":
        return _extract_text(file_bytes)
    else:
        raise ValueError(f"Unsupported file type: .{ext}")


# ── PDF ──────────────────────────────────────────────────────────────────────

def _extract_pdf(file_bytes: bytes) -> list:
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    text = ""
    for page in doc:
        text += page.get_text()

    if text.strip():
        # Text-based PDF — send extracted text, faster and cheaper
        return [{"type": "text", "text": text}]
    else:
        # Scanned/image PDF — send raw bytes to Claude vision
        return [{
            "type": "document",
            "source": {
                "type": "base64",
                "media_type": "application/pdf",
                "data": base64.standard_b64encode(file_bytes).decode()
            }
        }]


# ── DOCX ─────────────────────────────────────────────────────────────────────

def _extract_docx(file_bytes: bytes) -> list:
    # Build rId → (mime, image_bytes) map from the ZIP relationships
    rid_to_image = {}
    with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
        if "word/_rels/document.xml.rels" in z.namelist():
            rels_root = etree.fromstring(z.read("word/_rels/document.xml.rels"))
            for rel in rels_root:
                rid = rel.get("Id")
                target = rel.get("Target", "")
                if target.startswith("media/"):
                    full_path = f"word/{target}"
                    if full_path in z.namelist():
                        ext = target.rsplit(".", 1)[-1].lower()
                        mime = _image_mime(ext)
                        if mime:
                            rid_to_image[rid] = (mime, z.read(full_path))

    # Walk document XML in order — preserves text + image sequence
    doc = Document(io.BytesIO(file_bytes))
    blocks = []

    for element in doc.element.body:
        tag = element.tag.split("}")[-1] if "}" in element.tag else element.tag

        if tag == "p":
            blips = element.findall(".//" + qn("a:blip"))
            if blips:
                for blip in blips:
                    rid = blip.get(qn("r:embed"))
                    if rid and rid in rid_to_image:
                        mime, img_bytes = rid_to_image[rid]
                        blocks.append(_image_block(mime, img_bytes))
            else:
                text = "".join(
                    node.text or ""
                    for node in element.findall(".//" + qn("w:t"))
                )
                if text.strip():
                    blocks.append({"type": "text", "text": text})

        elif tag == "tbl":
            rows = []
            for row in element.findall(".//" + qn("w:tr")):
                cells = [
                    "".join(n.text or "" for n in cell.findall(".//" + qn("w:t")))
                    for cell in row.findall(".//" + qn("w:tc"))
                ]
                rows.append(" | ".join(cells))
            if rows:
                blocks.append({"type": "text", "text": "\n".join(rows)})

    return blocks


# ── XLSX ─────────────────────────────────────────────────────────────────────

def _extract_xlsx(file_bytes: bytes) -> list:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    blocks = []

    for sheet in wb.worksheets:
        blocks.append({"type": "text", "text": f"[Sheet: {sheet.title}]"})

        rows = []
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                rows.append(" | ".join(str(c) if c is not None else "" for c in row))
        if rows:
            blocks.append({"type": "text", "text": "\n".join(rows)})

        # Extract embedded images from xl/media/
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
            media = [f for f in z.namelist() if f.startswith("xl/media/")]
            for path in media:
                ext = path.rsplit(".", 1)[-1].lower()
                mime = _image_mime(ext)
                if mime:
                    blocks.append(_image_block(mime, z.read(path)))

    return blocks


# ── Image ─────────────────────────────────────────────────────────────────────

def _extract_image(file_bytes: bytes, ext: str) -> list:
    mime = _image_mime(ext)
    return [_image_block(mime, file_bytes)]


# ── Plain text ────────────────────────────────────────────────────────────────

def _extract_text(file_bytes: bytes) -> list:
    return [{"type": "text", "text": file_bytes.decode("utf-8", errors="replace")}]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _image_mime(ext: str) -> str | None:
    return {
        "png": "image/png",
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "gif": "image/gif",
        "webp": "image/webp",
    }.get(ext)


def _image_block(mime: str, img_bytes: bytes) -> dict:
    return {
        "type": "image",
        "source": {
            "type": "base64",
            "media_type": mime,
            "data": base64.standard_b64encode(img_bytes).decode()
        }
    }
